"""Shared pytest fixtures for RedWing DB Automation tests."""

from __future__ import annotations

import json
import os
import shutil
import sys
import tempfile
from pathlib import Path
from typing import Generator

import pytest
from fastapi.testclient import TestClient

# Add backend to path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from models import SubmissionPayload


# ── Paths ────────────────────────────────────────────────────────────────────

FIXTURES_DIR = Path(__file__).parent / "fixtures"
SAMPLE_WAYPOINTS_FILE = FIXTURES_DIR / "sample.waypoints"


# ── Sample Submission Payload ────────────────────────────────────────────────

SAMPLE_PAYLOAD_DICT = {
    "network_name": "Hoskote - Network Zero",
    "source_location_name": "HQ - Redwing Techworks",
    "source_takeoff_zone_name": "HQ North Pad",
    "source_latitude": 13.1637751,
    "source_longitude": 77.8672772,
    "destination_location_name": "Demo Site Alpha",
    "destination_landing_zone_name": "Demo Alpha South Pad",
    "destination_latitude": 13.2100000,
    "destination_longitude": 77.9100000,
    "takeoff_direction": 180,
    "approach_direction": 90,
    "mission_filename": "HQ-DEMO-180m.waypoints",
    "mission_drive_link": "https://drive.google.com/file/d/xxxxx",
    "elevation_image_drive_link": "https://drive.google.com/file/d/yyyyy",
    "route_image_drive_link": "https://drive.google.com/file/d/zzzzz",
}


@pytest.fixture
def sample_payload() -> SubmissionPayload:
    return SubmissionPayload(**SAMPLE_PAYLOAD_DICT)


@pytest.fixture
def sample_payload_dict() -> dict:
    return SAMPLE_PAYLOAD_DICT.copy()


# ── Temp Directory for Test Repo ─────────────────────────────────────────────

@pytest.fixture
def test_repo_path(tmp_path: Path) -> Path:
    """Create a temporary repo structure mirroring RedWingGCS."""
    repo = tmp_path / "RedWingGCS"
    repo.mkdir()
    (repo / "missions").mkdir()
    (repo / "frontend" / "public" / "Elevation and flight routes").mkdir(parents=True)
    (repo / "instance").mkdir()
    return repo


# ── Test Excel File ──────────────────────────────────────────────────────────

@pytest.fixture
def test_excel_path(test_repo_path: Path) -> Path:
    """Create a test Excel file with the correct sheet structure."""
    import openpyxl

    wb = openpyxl.Workbook()

    # Networks sheet
    ws_net = wb.active
    ws_net.title = "Networks"
    ws_net.append(["id", "name", "description"])
    ws_net.append([1, "Hoskote - Network Zero", "Test network"])
    ws_net.append([2, "Bangalore - Network One", "Another network"])

    # Locations sheet
    ws_loc = wb.create_sheet("Locations")
    ws_loc.append([
        "id", "name", "code", "district_id", "network_id",
        "location_type_id", "landing_zone_count",
    ])
    ws_loc.append([1, "HQ - Redwing Techworks", "HQ-RWT", 1, 1, 1, 2])
    ws_loc.append([2, "Depot - Main Warehouse", "D-MW", 1, 1, 1, 1])
    ws_loc.append([3, "Demo Site Alpha", "DSA", 1, 1, 1, 0])

    # Landing Zone sheet
    ws_lz = wb.create_sheet("Landing Zone")
    ws_lz.append([
        "id", "location_id", "name", "latitude", "longitude",
        "Status", "altitude_pixhawk", "altitude_srtm", "altitude_google_earth",
    ])
    ws_lz.append([1, 1, "HQ North Pad", 13.1637751, 77.8672772, 1, 0, 0, 0])
    ws_lz.append([2, 1, "HQ South Pad", 13.1635000, 77.8670000, 1, 0, 0, 0])
    ws_lz.append([3, 2, "Depot Pad A", 13.1800000, 77.8800000, 1, 0, 0, 0])

    # Waypoint Files sheet
    ws_wp = wb.create_sheet("Waypoint Files")
    ws_wp.append([
        "id", "filename", "google_drive_filelink",
        "local_filepath", "elevation_image", "route_image",
    ])
    ws_wp.append([
        1, "HQ-DEPOT-150m.waypoints",
        "https://drive.google.com/file/d/abc123",
        "./missions/HQ-DEPOT-150m.waypoints",
        "./Elevation and flight routes/HQ-DEPOT-150m elevation graph.png",
        "./Elevation and flight routes/HQ-DEPOT-150m flight route.png",
    ])

    # Flight Routes sheet
    ws_fr = wb.create_sheet("Flight Routes")
    ws_fr.append([
        "id", "start_lz_id", "end_lz_id", "start_location_id",
        "end_location_id", "waypoint_file_id", "network_id",
        "takeoff_direction", "approach_direction", "status",
        "distance", "flight_duration",
    ])
    ws_fr.append([1, 1, 3, 1, 2, 1, 1, 90, 270, 1, None, None])

    excel_path = test_repo_path / "Flight_data_updated.xlsx"
    wb.save(str(excel_path))
    wb.close()

    return excel_path


# ── Sample Waypoints File Fixture ────────────────────────────────────────────

@pytest.fixture
def sample_waypoints_path() -> Path:
    return SAMPLE_WAYPOINTS_FILE


@pytest.fixture
def temp_waypoints_file(tmp_path: Path) -> Path:
    """Copy sample waypoints to a temp location."""
    dest = tmp_path / "sample.waypoints"
    shutil.copy(str(SAMPLE_WAYPOINTS_FILE), str(dest))
    return dest


# ── FastAPI Test Client ──────────────────────────────────────────────────────

@pytest.fixture
def test_settings(test_repo_path: Path, test_excel_path: Path, tmp_path: Path):
    """Override settings for testing."""
    from config import Settings

    return Settings(
        REDWING_REPO_PATH=str(test_repo_path),
        EXCEL_FILENAME="Flight_data_updated.xlsx",
        POPULATE_SCRIPT="populate_data.py",
        GIT_BRANCH="main",
        SUBMISSIONS_DB_PATH=str(tmp_path / "test_submissions.db"),
        WEBHOOK_SECRET="test-secret",
        CESIUM_ION_TOKEN="test-token",
        CORS_ORIGINS="http://localhost:5173",
    )


@pytest.fixture
def test_client(test_settings, tmp_path: Path) -> Generator[TestClient, None, None]:
    """Create a FastAPI test client with test settings."""
    from config import get_settings
    from main import app, _store, get_store
    from submission_store import SubmissionStore
    import main as main_module

    # Override settings
    app.dependency_overrides[get_settings] = lambda: test_settings

    # Override store
    store = SubmissionStore(test_settings.SUBMISSIONS_DB_PATH)
    main_module._store = store
    app.dependency_overrides[get_store] = lambda: store

    with TestClient(app) as client:
        yield client

    app.dependency_overrides.clear()
    main_module._store = None


@pytest.fixture
def test_store(tmp_path: Path) -> SubmissionStore:
    """Create a test submission store."""
    from submission_store import SubmissionStore

    return SubmissionStore(str(tmp_path / "test.db"))
