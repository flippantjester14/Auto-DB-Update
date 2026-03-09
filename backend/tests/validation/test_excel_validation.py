"""Post-approval validation: verify the updated Excel file."""

from __future__ import annotations

import os
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "../.."))


@pytest.fixture
def approved_excel(test_excel_path: Path, sample_payload) -> Path:
    from excel_updater import ExcelUpdater
    updater = ExcelUpdater(test_excel_path)
    updater.open()
    updater.execute_pipeline(sample_payload)
    updater.close()
    return test_excel_path


class TestFlightRouteIntegrity:
    def test_new_route_exists(self, approved_excel: Path):
        wb = openpyxl.load_workbook(str(approved_excel))
        ws = wb["Flight Routes"]
        assert ws.max_row >= 3
        assert ws.cell(row=ws.max_row, column=1).value == 2
        wb.close()

    def test_route_references_valid_ids(self, approved_excel: Path):
        wb = openpyxl.load_workbook(str(approved_excel))
        ws_fr = wb["Flight Routes"]
        ws_lz = wb["Landing Zone"]
        ws_loc = wb["Locations"]
        ws_wp = wb["Waypoint Files"]
        ws_net = wb["Networks"]
        last = ws_fr.max_row
        fr_h = {ws_fr.cell(1, c).value.strip(): c for c in range(1, ws_fr.max_column+1) if ws_fr.cell(1,c).value}
        lz_ids = {ws_lz.cell(r,1).value for r in range(2, ws_lz.max_row+1)}
        loc_ids = {ws_loc.cell(r,1).value for r in range(2, ws_loc.max_row+1)}
        wp_ids = {ws_wp.cell(r,1).value for r in range(2, ws_wp.max_row+1)}
        net_ids = {ws_net.cell(r,1).value for r in range(2, ws_net.max_row+1)}
        assert ws_fr.cell(last, fr_h["start_lz_id"]).value in lz_ids
        assert ws_fr.cell(last, fr_h["end_lz_id"]).value in lz_ids
        assert ws_fr.cell(last, fr_h["start_location_id"]).value in loc_ids
        assert ws_fr.cell(last, fr_h["end_location_id"]).value in loc_ids
        assert ws_fr.cell(last, fr_h["waypoint_file_id"]).value in wp_ids
        assert ws_fr.cell(last, fr_h["network_id"]).value in net_ids
        wb.close()


class TestNoDuplicateIDs:
    @pytest.mark.parametrize("sheet_name", [
        "Networks", "Locations", "Landing Zone", "Waypoint Files", "Flight Routes",
    ])
    def test_no_duplicate_ids(self, approved_excel: Path, sheet_name: str):
        wb = openpyxl.load_workbook(str(approved_excel))
        ws = wb[sheet_name]
        ids = [ws.cell(r,1).value for r in range(2, ws.max_row+1) if ws.cell(r,1).value is not None]
        assert len(ids) == len(set(ids)), f"Duplicate IDs in {sheet_name}"
        wb.close()


class TestWaypointFilePaths:
    def test_paths_format(self, approved_excel: Path):
        wb = openpyxl.load_workbook(str(approved_excel))
        ws = wb["Waypoint Files"]
        h = {ws.cell(1,c).value.strip(): c for c in range(1, ws.max_column+1) if ws.cell(1,c).value}
        for row in range(2, ws.max_row+1):
            fn = ws.cell(row, h["filename"]).value
            if fn is None: continue
            assert ws.cell(row, h["local_filepath"]).value.startswith("./missions/")
            assert "elevation graph.png" in ws.cell(row, h["elevation_image"]).value
            assert "flight route.png" in ws.cell(row, h["route_image"]).value
        wb.close()
