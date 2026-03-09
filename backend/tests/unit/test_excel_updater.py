"""Unit tests for excel_updater.py."""

from __future__ import annotations

import os
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "../.."))

from excel_updater import ExcelUpdater, _auto_generate_code
from models import EntityAction, SubmissionPayload


class TestAutoGenerateCode:
    """Test location code auto-generation."""

    def test_simple_name(self):
        assert _auto_generate_code("Demo Site Alpha") == "DSA"

    def test_dash_separated(self):
        assert _auto_generate_code("HQ - Redwing Techworks") == "H-RT"

    def test_max_length(self):
        code = _auto_generate_code("A Very Long Location Name Here Exceeding Limit")
        assert len(code) <= 8

    def test_single_word(self):
        assert _auto_generate_code("Headquarters") == "H"


class TestResolveNetwork:
    """Test network resolution."""

    def test_existing_network_found(self, test_excel_path: Path):
        updater = ExcelUpdater(test_excel_path)
        updater.open()
        preview, nid = updater.resolve_network("Hoskote - Network Zero")
        assert preview.action == EntityAction.EXISTING
        assert nid == 1
        assert preview.id == 1
        updater.close()

    def test_network_not_found(self, test_excel_path: Path):
        updater = ExcelUpdater(test_excel_path)
        updater.open()
        preview, nid = updater.resolve_network("Nonexistent Network")
        assert preview.action == EntityAction.NOT_FOUND
        assert nid is None
        updater.close()

    def test_network_not_found_raises_in_pipeline(self, test_excel_path: Path, sample_payload: SubmissionPayload):
        updater = ExcelUpdater(test_excel_path)
        updater.open()
        # Modify payload to use nonexistent network
        payload = sample_payload.model_copy(update={"network_name": "Fake Network"})
        with pytest.raises(RuntimeError, match="Network not found"):
            updater.execute_pipeline(payload)
        updater.close()


class TestResolveLocation:
    """Test location resolution."""

    def test_existing_location_found(self, test_excel_path: Path):
        updater = ExcelUpdater(test_excel_path)
        updater.open()
        preview, loc_id = updater.resolve_location("HQ - Redwing Techworks", 1)
        assert preview.action == EntityAction.EXISTING
        assert loc_id == 1
        updater.close()

    def test_new_location_created(self, test_excel_path: Path):
        updater = ExcelUpdater(test_excel_path)
        updater.open()
        preview, loc_id = updater.resolve_location("Brand New Location", 1)
        assert preview.action == EntityAction.NEW
        assert loc_id == 4  # 3 existing + 1 new
        updater.close()

    def test_new_location_has_correct_code(self, test_excel_path: Path):
        updater = ExcelUpdater(test_excel_path)
        updater.open()
        updater.resolve_location("Test - Medical Center", 1)

        # Read back from Excel
        ws = updater._ws("Locations")
        last_row = ws.max_row
        code = ws.cell(row=last_row, column=3).value  # code column
        assert code is not None
        assert code == "T-MC"
        updater.close()

    def test_id_incremented_correctly(self, test_excel_path: Path):
        updater = ExcelUpdater(test_excel_path)
        updater.open()
        _, id1 = updater.resolve_location("Location A", 1)
        _, id2 = updater.resolve_location("Location B", 1)
        assert id2 == id1 + 1
        updater.close()

    def test_dry_run_does_not_write(self, test_excel_path: Path):
        updater = ExcelUpdater(test_excel_path)
        updater.open()
        ws = updater._ws("Locations")
        rows_before = ws.max_row

        preview, _ = updater.resolve_location("Dry Run Location", 1, dry_run=True)
        assert preview.action == EntityAction.NEW
        assert ws.max_row == rows_before  # no new row
        updater.close()


class TestResolveLandingZone:
    """Test landing zone resolution."""

    def test_existing_lz_found(self, test_excel_path: Path):
        updater = ExcelUpdater(test_excel_path)
        updater.open()
        preview, lz_id = updater.resolve_landing_zone(
            "HQ North Pad", 1, 13.1637751, 77.8672772
        )
        assert preview.action == EntityAction.EXISTING
        assert lz_id == 1
        updater.close()

    def test_lat_long_mismatch_creates_new(self, test_excel_path: Path):
        updater = ExcelUpdater(test_excel_path)
        updater.open()
        preview, lz_id = updater.resolve_landing_zone(
            "HQ North Pad", 1, 99.0, 99.0  # totally different coords
        )
        assert preview.action == EntityAction.NEW
        assert lz_id == 4  # 3 existing + 1 new
        updater.close()

    def test_new_lz_increments_parent_count(self, test_excel_path: Path):
        updater = ExcelUpdater(test_excel_path)
        updater.open()

        # Location 3 (Demo Site Alpha) has landing_zone_count = 0
        ws_loc = updater._ws("Locations")
        headers = updater._headers("Locations")
        count_col = headers["landing_zone_count"]
        assert ws_loc.cell(row=4, column=count_col).value == 0

        updater.resolve_landing_zone(
            "New Pad at Demo", 3, 13.21, 77.91
        )

        assert ws_loc.cell(row=4, column=count_col).value == 1
        updater.close()

    def test_dry_run_does_not_write(self, test_excel_path: Path):
        updater = ExcelUpdater(test_excel_path)
        updater.open()
        ws = updater._ws("Landing Zone")
        rows_before = ws.max_row

        preview, _ = updater.resolve_landing_zone(
            "Dry Run Pad", 1, 99.0, 99.0, dry_run=True
        )
        assert preview.action == EntityAction.NEW
        assert ws.max_row == rows_before
        updater.close()


class TestRegisterWaypointFile:
    """Test waypoint file registration."""

    def test_always_creates_new(self, test_excel_path: Path):
        updater = ExcelUpdater(test_excel_path)
        updater.open()
        preview, wp_id = updater.register_waypoint_file(
            "test.waypoints", "https://drive.google.com/file/d/test"
        )
        assert preview.action == EntityAction.NEW
        assert wp_id == 2  # 1 existing + 1 new
        updater.close()

    def test_correct_paths_constructed(self, test_excel_path: Path):
        updater = ExcelUpdater(test_excel_path)
        updater.open()
        updater.register_waypoint_file(
            "HQ-DEMO-180m.waypoints", "https://drive.google.com/file/d/test"
        )

        ws = updater._ws("Waypoint Files")
        last_row = ws.max_row
        headers = updater._headers("Waypoint Files")

        filepath = ws.cell(row=last_row, column=headers["local_filepath"]).value
        assert filepath == "./missions/HQ-DEMO-180m.waypoints"

        elev = ws.cell(row=last_row, column=headers["elevation_image"]).value
        assert elev == "./Elevation and flight routes/HQ-DEMO-180m elevation graph.png"

        route_img = ws.cell(row=last_row, column=headers["route_image"]).value
        assert route_img == "./Elevation and flight routes/HQ-DEMO-180m flight route.png"
        updater.close()


class TestCreateFlightRoute:
    """Test flight route creation."""

    def test_creates_with_correct_ids(self, test_excel_path: Path):
        updater = ExcelUpdater(test_excel_path)
        updater.open()
        preview, route_id = updater.create_flight_route(
            start_lz_id=1,
            end_lz_id=3,
            start_location_id=1,
            end_location_id=2,
            waypoint_file_id=1,
            network_id=1,
            takeoff_direction=180,
            approach_direction=90,
        )
        assert preview.action == EntityAction.NEW
        assert route_id == 2  # 1 existing + 1 new

        ws = updater._ws("Flight Routes")
        last_row = ws.max_row
        headers = updater._headers("Flight Routes")
        assert ws.cell(row=last_row, column=headers["start_lz_id"]).value == 1
        assert ws.cell(row=last_row, column=headers["end_lz_id"]).value == 3
        assert ws.cell(row=last_row, column=headers["takeoff_direction"]).value == 180
        assert ws.cell(row=last_row, column=headers["approach_direction"]).value == 90
        assert ws.cell(row=last_row, column=headers["status"]).value == 1
        updater.close()


class TestIDIntegrity:
    """Test sequential ID integrity across multiple operations."""

    def test_no_gaps_or_duplicates(self, test_excel_path: Path):
        updater = ExcelUpdater(test_excel_path)
        updater.open()

        # Add multiple locations
        ids = []
        for i in range(5):
            _, loc_id = updater.resolve_location(f"Test Location {i}", 1)
            ids.append(loc_id)

        # Verify sequential, no gaps
        for i in range(1, len(ids)):
            assert ids[i] == ids[i - 1] + 1

        # Verify no duplicates
        assert len(set(ids)) == len(ids)

        updater.close()

    def test_full_pipeline_ids_consistent(self, test_excel_path: Path, sample_payload: SubmissionPayload):
        updater = ExcelUpdater(test_excel_path)
        updater.open()
        ids = updater.execute_pipeline(sample_payload)
        updater.close()

        # Verify all IDs are present and non-None
        for key in [
            "network_id", "source_location_id", "source_lz_id",
            "destination_location_id", "destination_lz_id",
            "waypoint_file_id", "flight_route_id",
        ]:
            assert ids[key] is not None
            assert isinstance(ids[key], int)
            assert ids[key] > 0

        # Verify the saved Excel can be re-parsed
        updater2 = ExcelUpdater(test_excel_path)
        updater2.open()
        # The flight route should now exist
        ws = updater2._ws("Flight Routes")
        last_id = ws.cell(row=ws.max_row, column=1).value
        assert last_id == ids["flight_route_id"]
        updater2.close()
